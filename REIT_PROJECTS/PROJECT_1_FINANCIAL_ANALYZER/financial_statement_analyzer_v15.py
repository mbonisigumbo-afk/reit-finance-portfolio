import pandas as pd
from pathlib import Path
import matplotlib.pyplot as plt
from datetime import datetime

from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image


def load_data():

    file_path = (
        Path(__file__).parent
        / "reit_financials_real.xlsx"
    )

    try:
        df = pd.read_excel(file_path)
        return df

    except FileNotFoundError:
        print("Financial file not found.")
        exit()


def validate_data(df):

    if (df["Revenue"] < 0).any():
        print("WARNING: Negative revenue detected")

    if (df["Total_Assets"] <= 0).any():
        print("WARNING: Invalid asset values detected")


def calculate_ratios(df):

    df["ROA"] = (
        df["Net_Income"]
        / df["Total_Assets"]
    ) * 100

    df["Debt_to_Assets"] = (
        df["Debt"]
        / df["Total_Assets"]
    ) * 100

    df["Revenue_per_Asset"] = (
        df["Revenue"]
        / df["Total_Assets"]
    ) * 100

    return df


def debt_risk_rating(debt_ratio):

    if debt_ratio < 35:
        return "LOW RISK"

    elif debt_ratio < 50:
        return "NORMAL"

    else:
        return "HIGH RISK"


def generate_summary(df):

    print("\n=== REIT ANALYST SUMMARY ===\n")

    for _, row in df.iterrows():

        print(
            f"Rank #{row['ROA_Rank']} | "
            f"{row['REIT']} | "
            f"ROA={row['ROA']:.2f}% | "
            f"Debt={row['Debt_to_Assets']:.2f}% | "
            f"Risk={row['Risk_Rating']}"
        )


def portfolio_statistics(df):

    print("\n=== PORTFOLIO STATISTICS ===\n")

    print(
        f"Average ROA: "
        f"{df['ROA'].mean():.2f}%"
    )

    print(
        f"Average Debt Ratio: "
        f"{df['Debt_to_Assets'].mean():.2f}%"
    )

    print(
        f"Total Assets: "
        f"R{df['Total_Assets'].sum():,.0f}"
    )


def create_roa_chart(df):

    plt.figure()

    plt.bar(
        df["REIT"],
        df["ROA"]
    )

    plt.title("ROA by REIT")
    plt.ylabel("ROA (%)")

    plt.savefig(
        Path(__file__).parent
        / "roa_chart.png"
    )

    plt.close()


def create_debt_chart(df):

    plt.figure()

    plt.bar(
        df["REIT"],
        df["Debt_to_Assets"]
    )

    plt.title("Debt-to-Assets by REIT")
    plt.ylabel("Debt Ratio (%)")

    plt.savefig(
        Path(__file__).parent
        / "debt_chart.png"
    )

    plt.close()


def create_revenue_efficiency_chart(df):

    plt.figure()

    plt.bar(
        df["REIT"],
        df["Revenue_per_Asset"]
    )

    plt.title("Revenue per Asset by REIT")
    plt.ylabel("Revenue per Asset (%)")

    plt.savefig(
        Path(__file__).parent
        / "revenue_efficiency_chart.png"
    )

    plt.close()


def auto_fit_columns(sheet):

    for column in sheet.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:

                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            except Exception:
                pass

        sheet.column_dimensions[
            column_letter
        ].width = max_length + 2


def main():

    df = load_data()

    validate_data(df)

    df = calculate_ratios(df)

    df["Risk_Rating"] = (
        df["Debt_to_Assets"]
        .apply(debt_risk_rating)
    )

    df["ROA_Rank"] = (
        df["ROA"]
        .rank(
            ascending=False,
            method="dense"
        )
        .astype(int)
    )

    df = df.sort_values(
        by="ROA",
        ascending=False
    )

    generate_summary(df)

    best_reit = df.loc[
        df["ROA"].idxmax()
    ]

    print("\n=== TOP PERFORMER ===\n")

    print(
        f"{best_reit['REIT']} "
        f"has the highest ROA at "
        f"{best_reit['ROA']:.2f}%"
    )

    portfolio_statistics(df)

    print("\n=== DASHBOARD METRICS ===\n")

    print(
        f"Highest ROA: "
        f"{df.loc[df['ROA'].idxmax(), 'REIT']}"
    )

    print(
        f"Lowest Debt Ratio: "
        f"{df.loc[df['Debt_to_Assets'].idxmin(), 'REIT']}"
    )

    print(
        f"Best Revenue Efficiency: "
        f"{df.loc[df['Revenue_per_Asset'].idxmax(), 'REIT']}"
    )

    create_roa_chart(df)
    create_debt_chart(df)
    create_revenue_efficiency_chart(df)

    dashboard_df = pd.DataFrame({

    "Metric": [

        "Report Date",
        "Number of REITs Analysed",
        "Highest ROA",
        "Average ROA",
        "Lowest Debt Ratio",
        "Average Debt Ratio",
        "Best Revenue Efficiency",
        "Largest REIT (Assets)",
        "Smallest REIT (Assets)",
        "Total Assets"

    ],

    "Value": [

        datetime.now().strftime("%Y-%m-%d"),

        len(df),

        df.loc[
            df["ROA"].idxmax(),
            "REIT"
        ],

        f"{df['ROA'].mean():.2f}%",

        df.loc[
            df["Debt_to_Assets"].idxmin(),
            "REIT"
        ],

        f"{df['Debt_to_Assets'].mean():.2f}%",

        df.loc[
            df["Revenue_per_Asset"].idxmax(),
            "REIT"
        ],

        df.loc[
            df["Total_Assets"].idxmax(),
            "REIT"
        ],

        df.loc[
            df["Total_Assets"].idxmin(),
            "REIT"
        ],

        f"R{df['Total_Assets'].sum():,.0f}"

    ]

})

    alerts = []

    for _, row in df.iterrows():

        if row["ROA"] >= 5:
            alerts.append("Strong ROA")

        elif row["ROA"] >= 2:
            alerts.append("Average ROA")

        else:
            alerts.append("Weak ROA")

    alerts_df = pd.DataFrame({

        "REIT": df["REIT"],

        "ROA": df["ROA"].round(2),

        "Debt Ratio": df["Debt_to_Assets"].round(2),

        "Alert": alerts

    })

    output_file = (
        Path(__file__).parent
        / "reit_analysis_output.xlsx"
    )

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name="Analysis",
            index=False
        )

        dashboard_df.to_excel(
            writer,
            sheet_name="Dashboard",
            index=False
        )

        alerts_df.to_excel(
            writer,
            sheet_name="Alerts",
            index=False
        )

        workbook = writer.book

        workbook.create_sheet("Charts")

        analysis_sheet = workbook["Analysis"]
        dashboard_sheet = workbook["Dashboard"]
        alerts_sheet = workbook["Alerts"]
        charts_sheet = workbook["Charts"]

        green_fill = PatternFill(
            fill_type="solid",
            start_color="C6EFCE"
        )

        yellow_fill = PatternFill(
            fill_type="solid",
            start_color="FFF2CC"
        )

        red_fill = PatternFill(
            fill_type="solid",
            start_color="F4CCCC"
        )

        # Format alert cells
        for row in range(
            2,
            alerts_sheet.max_row + 1
        ):

            cell = alerts_sheet[f"D{row}"]

            if cell.value == "Strong ROA":

                cell.fill = green_fill

            elif cell.value == "Average ROA":

                cell.fill = yellow_fill

            elif cell.value == "Weak ROA":

                cell.fill = red_fill

        # Bold headers
        for cell in analysis_sheet[1]:

            cell.font = Font(
                bold=True
            )

        # Insert title row
        dashboard_sheet.insert_rows(1)

        dashboard_sheet["A1"] = "REIT EXECUTIVE DASHBOARD"

        dashboard_sheet["A1"].font = Font(
                bold=True,
                size=16
        )

        # Format Dashboard header
        for cell in dashboard_sheet[2]:

                cell.font = Font(
                    bold=True,
                    size=12
                )

        # Set Dashboard column widths
        dashboard_sheet.column_dimensions["A"].width = 32
        dashboard_sheet.column_dimensions["B"].width = 28
        # Dashboard metric highlighting

        highlight_fill = PatternFill(
            fill_type="solid",
            start_color="E2F0D9"
        )

        for row in range(
            3,    
            dashboard_sheet.max_row + 1 
        ):

            metric_cell = dashboard_sheet[f"A{row}"]
            value_cell = dashboard_sheet[f"B{row}"]

            if metric_cell.value in [
                "Highest ROA",
                "Lowest Debt Ratio",
                "Best Revenue Efficiency"
            ]:

               metric_cell.font = Font(
                   bold=True
             )

               value_cell.font = Font(
                   bold=True
               )

               metric_cell.fill = highlight_fill
               value_cell.fill = highlight_fill

        # Freeze Dashboard header
        dashboard_sheet.freeze_panes = "A3"
        dashboard_sheet["A1"] = "REIT EXECUTIVE DASHBOARD"

        dashboard_sheet["A1"].font = Font(
                bold=True,
                size=14
)

        for cell in alerts_sheet[1]:

            cell.font = Font(
                bold=True
            )

        # Auto-fit columns
        auto_fit_columns(
            analysis_sheet
        )

        auto_fit_columns(
            dashboard_sheet
        )

        auto_fit_columns(
            alerts_sheet
        )

        # Load chart images
        roa_image = Image(
            str(
                Path(__file__).parent
                / "roa_chart.png"
            )
        )

        debt_image = Image(
            str(
                Path(__file__).parent
                / "debt_chart.png"
            )
        )

        revenue_image = Image(
            str(
                Path(__file__).parent
                / "revenue_efficiency_chart.png"
            )
        )

        # Add charts to Charts sheet
        charts_sheet.add_image(
            roa_image,
            "A1"
        )

        charts_sheet.add_image(
            debt_image,
            "A20"
        )

        charts_sheet.add_image(
            revenue_image,
            "A40"
        )

    print("\nExcel report created.")


if __name__ == "__main__":
    main()